"""
가공 견적 AI 엔진 PRO v2.3
main.py — Streamlit 메인 UI

변경사항 (v2.2 → v2.3):
- [UI] 전체 커스텀 CSS 적용 — 카드형 레이아웃, 색상 통일, 폰트 개선
- [신규] 일괄 업로드 — STEP + 명세서 동시, 형번 자동 매칭
- [신규] 업체별 패턴 분석 — 난이도별 마진 포인트, 추이 분석
- [신규] 시세 마지막 갱신 시각 표시 + 3일 미갱신 경고
- [유지] 역산 감사, 업체 시뮬레이션, 협상 근거 엑셀, 유사 부품 매칭
"""

import io
import sqlite3
import tempfile
import os
import json
from datetime import datetime, date, timedelta
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from engine import (
    audit, audit_scenario, AuditResult,
    calc_material_cost, load_material_master, get_market_prices,
    get_coefficient, get_label, get_all_options, parse_level_from_option,
    DIFFICULTY_TABLE,
    analyze_step_file, is_step_available,
)

BASE_DIR = Path(__file__).parent
DB_PATH  = BASE_DIR / "data" / "history_log.db"


# ══════════════════════════════════════════════════════════════
# CSS 커스텀 스타일
# ══════════════════════════════════════════════════════════════
def inject_css():
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

    html, body, [class*="css"] {
        font-family: 'Noto Sans KR', sans-serif !important;
    }

    /* 전체 배경 */
    .stApp {
        background-color: #F0F2F6;
    }

    /* 사이드바 */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1B2A4A 0%, #243B6E 100%) !important;
        border-right: none !important;
    }
    section[data-testid="stSidebar"] * {
        color: #E8EDF5 !important;
    }
    section[data-testid="stSidebar"] .stRadio label {
        padding: 8px 12px !important;
        border-radius: 8px !important;
        cursor: pointer !important;
        transition: background .15s !important;
        font-size: 14px !important;
    }
    section[data-testid="stSidebar"] .stRadio label:hover {
        background: rgba(255,255,255,0.1) !important;
    }
    section[data-testid="stSidebar"] .stSelectbox select,
    section[data-testid="stSidebar"] .stNumberInput input {
        background: rgba(255,255,255,0.1) !important;
        border: 1px solid rgba(255,255,255,0.2) !important;
        color: #E8EDF5 !important;
        border-radius: 6px !important;
    }
    section[data-testid="stSidebar"] .stSlider > div > div {
        background: rgba(255,255,255,0.2) !important;
    }
    section[data-testid="stSidebar"] hr {
        border-color: rgba(255,255,255,0.15) !important;
    }
    section[data-testid="stSidebar"] .stButton button {
        background: rgba(255,255,255,0.15) !important;
        border: 1px solid rgba(255,255,255,0.25) !important;
        color: #E8EDF5 !important;
        border-radius: 6px !important;
        font-size: 12px !important;
    }

    /* 메인 컨텐츠 */
    .main .block-container {
        padding: 2rem 2.5rem !important;
        max-width: 1400px !important;
    }

    /* 페이지 제목 */
    h1 {
        font-size: 1.6rem !important;
        font-weight: 700 !important;
        color: #1B2A4A !important;
        margin-bottom: 1.5rem !important;
        padding-bottom: 0.75rem !important;
        border-bottom: 2px solid #3B82F6 !important;
    }

    /* 섹션 제목 */
    h2, h3 {
        font-size: 1rem !important;
        font-weight: 600 !important;
        color: #1B2A4A !important;
        margin-top: 1rem !important;
    }

    /* 카드 컨테이너 */
    div[data-testid="stVerticalBlock"] > div[data-testid="stVerticalBlock"] {
        background: #FFFFFF;
        border-radius: 12px;
        padding: 1.25rem;
        border: 1px solid #E2E8F0;
        box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        margin-bottom: 1rem;
    }

    /* 입력 필드 */
    .stTextInput input, .stNumberInput input, .stSelectbox select, .stTextArea textarea {
        border: 1.5px solid #E2E8F0 !important;
        border-radius: 8px !important;
        font-family: 'Noto Sans KR', sans-serif !important;
        font-size: 13px !important;
        transition: border-color .15s !important;
        background: #FAFBFC !important;
    }
    .stTextInput input:focus, .stNumberInput input:focus,
    .stSelectbox select:focus, .stTextArea textarea:focus {
        border-color: #3B82F6 !important;
        box-shadow: 0 0 0 3px rgba(59,130,246,0.1) !important;
        background: #FFFFFF !important;
    }

    /* 기본 버튼 */
    .stButton button {
        border-radius: 8px !important;
        font-family: 'Noto Sans KR', sans-serif !important;
        font-weight: 500 !important;
        font-size: 13px !important;
        transition: all .15s !important;
        border: 1.5px solid #E2E8F0 !important;
    }
    .stButton button:hover {
        border-color: #3B82F6 !important;
        color: #3B82F6 !important;
        transform: translateY(-1px) !important;
        box-shadow: 0 2px 8px rgba(59,130,246,0.15) !important;
    }

    /* Primary 버튼 */
    .stButton button[kind="primary"] {
        background: linear-gradient(135deg, #2563EB, #3B82F6) !important;
        color: white !important;
        border: none !important;
        box-shadow: 0 2px 8px rgba(37,99,235,0.3) !important;
    }
    .stButton button[kind="primary"]:hover {
        background: linear-gradient(135deg, #1D4ED8, #2563EB) !important;
        color: white !important;
        box-shadow: 0 4px 12px rgba(37,99,235,0.4) !important;
    }

    /* 다운로드 버튼 */
    .stDownloadButton button {
        background: linear-gradient(135deg, #059669, #10B981) !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        font-weight: 500 !important;
        box-shadow: 0 2px 8px rgba(5,150,105,0.3) !important;
    }
    .stDownloadButton button:hover {
        background: linear-gradient(135deg, #047857, #059669) !important;
        color: white !important;
        transform: translateY(-1px) !important;
        box-shadow: 0 4px 12px rgba(5,150,105,0.4) !important;
    }

    /* 메트릭 카드 */
    div[data-testid="metric-container"] {
        background: #F8FAFF !important;
        border: 1.5px solid #DBEAFE !important;
        border-radius: 10px !important;
        padding: 1rem !important;
    }
    div[data-testid="metric-container"] label {
        color: #64748B !important;
        font-size: 12px !important;
        font-weight: 500 !important;
    }
    div[data-testid="metric-container"] div[data-testid="stMetricValue"] {
        color: #1B2A4A !important;
        font-size: 1.3rem !important;
        font-weight: 700 !important;
    }

    /* 데이터프레임 */
    .stDataFrame {
        border-radius: 10px !important;
        overflow: hidden !important;
        border: 1px solid #E2E8F0 !important;
    }

    /* 알림 박스 */
    .stSuccess, .stWarning, .stError, .stInfo {
        border-radius: 10px !important;
        font-size: 13px !important;
    }

    /* 파일 업로더 */
    .stFileUploader {
        border: 2px dashed #CBD5E1 !important;
        border-radius: 10px !important;
        padding: 1rem !important;
        background: #F8FAFF !important;
        transition: border-color .15s !important;
    }
    .stFileUploader:hover {
        border-color: #3B82F6 !important;
        background: #EFF6FF !important;
    }

    /* expander */
    .streamlit-expanderHeader {
        background: #F8FAFF !important;
        border-radius: 8px !important;
        font-weight: 500 !important;
        font-size: 13px !important;
        color: #1B2A4A !important;
    }

    /* 탭 */
    .stTabs [data-baseweb="tab-list"] {
        gap: 4px !important;
        background: #F0F2F6 !important;
        padding: 4px !important;
        border-radius: 10px !important;
    }
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px !important;
        font-size: 13px !important;
        font-weight: 500 !important;
        padding: 6px 16px !important;
    }
    .stTabs [aria-selected="true"] {
        background: white !important;
        color: #2563EB !important;
        box-shadow: 0 1px 3px rgba(0,0,0,0.1) !important;
    }

    /* 구분선 */
    hr {
        border-color: #E2E8F0 !important;
        margin: 1rem 0 !important;
    }

    /* 캡션 */
    .stCaption, small {
        color: #94A3B8 !important;
        font-size: 11px !important;
    }

    /* 판정 배지 스타일 */
    .verdict-badge {
        display: inline-block;
        padding: 4px 12px;
        border-radius: 99px;
        font-size: 12px;
        font-weight: 600;
    }

    /* 숨김 라벨 */
    .stRadio > label, .stCheckbox > label {
        font-size: 13px !important;
        color: #374151 !important;
        font-weight: 500 !important;
    }
    </style>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
# DB 초기화
# ══════════════════════════════════════════════════════════════
def init_db():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS vendors (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                name TEXT NOT NULL, contact TEXT, phone TEXT,
                email TEXT, note TEXT, is_active INTEGER DEFAULT 1
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS history_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                part_name TEXT, part_no TEXT, step_file TEXT,
                material_code TEXT, vendor_id INTEGER REFERENCES vendors(id),
                vendor_name TEXT, vendor_price REAL, material_cost REAL,
                actual_machining REAL, standard_machining REAL,
                variance_pct REAL, verdict TEXT, difficulty_level INTEGER,
                hourly_rate REAL, estimated_hours REAL, postprocess_cost REAL,
                volume_cm3 REAL, hole_count INTEGER, setup_count INTEGER,
                price_snapshot REAL, ordered INTEGER DEFAULT 0, note TEXT
            )
        """)
        conn.commit()
    finally:
        conn.close()


# ══════════════════════════════════════════════════════════════
# 업체 CRUD
# ══════════════════════════════════════════════════════════════
def load_vendors(active_only=True) -> pd.DataFrame:
    conn = sqlite3.connect(DB_PATH)
    try:
        q = "SELECT * FROM vendors" + (" WHERE is_active=1" if active_only else "")
        return pd.read_sql(q + " ORDER BY name", conn)
    finally:
        conn.close()

def add_vendor(name, contact="", phone="", email="", note=""):
    conn = sqlite3.connect(DB_PATH)
    try:
        conn.execute("INSERT INTO vendors (name,contact,phone,email,note) VALUES (?,?,?,?,?)",
                     (name, contact, phone, email, note))
        conn.commit()
    finally:
        conn.close()

def update_vendor(vid, name, contact, phone, email, note):
    conn = sqlite3.connect(DB_PATH)
    try:
        conn.execute("UPDATE vendors SET name=?,contact=?,phone=?,email=?,note=? WHERE id=?",
                     (name, contact, phone, email, note, vid))
        conn.commit()
    finally:
        conn.close()

def deactivate_vendor(vid):
    conn = sqlite3.connect(DB_PATH)
    try:
        conn.execute("UPDATE vendors SET is_active=0 WHERE id=?", (vid,))
        conn.commit()
    finally:
        conn.close()

def vendor_has_history(vid) -> bool:
    conn = sqlite3.connect(DB_PATH)
    try:
        return conn.execute("SELECT COUNT(*) FROM history_log WHERE vendor_id=?",
                            (vid,)).fetchone()[0] > 0
    finally:
        conn.close()


# ══════════════════════════════════════════════════════════════
# 이력 저장·조회
# ══════════════════════════════════════════════════════════════
def save_to_db(part_name, part_no, step_file, material_code,
               vendor_id, vendor_name, result: AuditResult,
               difficulty_level, estimated_hours,
               volume_cm3=None, hole_count=None, setup_count=None,
               price_snapshot=None, note=""):
    conn = sqlite3.connect(DB_PATH)
    try:
        conn.execute("""
            INSERT INTO history_log (
                created_at,part_name,part_no,step_file,material_code,
                vendor_id,vendor_name,vendor_price,material_cost,
                actual_machining,standard_machining,variance_pct,verdict,
                difficulty_level,hourly_rate,estimated_hours,postprocess_cost,
                volume_cm3,hole_count,setup_count,price_snapshot,note
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (datetime.now().isoformat(), part_name, part_no, step_file, material_code,
              vendor_id, vendor_name, result.vendor_price, result.material_cost,
              result.actual_machining, result.standard_machining, result.variance_pct,
              result.verdict, difficulty_level, result.hourly_rate, estimated_hours,
              result.postprocess_cost, volume_cm3, hole_count, setup_count,
              price_snapshot, note))
        conn.commit()
    finally:
        conn.close()

def load_history() -> pd.DataFrame:
    if not DB_PATH.exists():
        return pd.DataFrame()
    conn = sqlite3.connect(DB_PATH)
    try:
        return pd.read_sql("SELECT * FROM history_log ORDER BY created_at DESC", conn)
    finally:
        conn.close()

def load_vendor_history(vendor_id) -> pd.DataFrame:
    conn = sqlite3.connect(DB_PATH)
    try:
        return pd.read_sql(
            "SELECT * FROM history_log WHERE vendor_id=? ORDER BY created_at DESC",
            conn, params=(vendor_id,))
    finally:
        conn.close()

def load_vendor_stats() -> pd.DataFrame:
    if not DB_PATH.exists():
        return pd.DataFrame()
    conn = sqlite3.connect(DB_PATH)
    try:
        try:
            return pd.read_sql("""
                SELECT v.id, v.name,
                    COUNT(h.id) AS total,
                    ROUND(AVG(h.variance_pct),1) AS avg_variance,
                    ROUND(MIN(h.variance_pct),1) AS min_variance,
                    ROUND(MAX(h.variance_pct),1) AS max_variance,
                    SUM(COALESCE(h.ordered,0)) AS ordered_count,
                    ROUND(AVG(h.material_cost/NULLIF(h.vendor_price,0)*100),1) AS mat_ratio,
                    ROUND(AVG(h.actual_machining/NULLIF(h.standard_machining,0)),3) AS avg_margin_ratio
                FROM vendors v
                LEFT JOIN history_log h ON v.id=h.vendor_id
                WHERE v.is_active=1
                GROUP BY v.id
            """, conn)
        except Exception:
            return pd.DataFrame()
    finally:
        conn.close()

def load_vendor_pattern(vendor_id) -> pd.DataFrame:
    """업체별 난이도별 평균 오차율 패턴 분석."""
    if not DB_PATH.exists():
        return pd.DataFrame()
    conn = sqlite3.connect(DB_PATH)
    try:
        return pd.read_sql("""
            SELECT
                difficulty_level,
                COUNT(*) AS cnt,
                ROUND(AVG(variance_pct),1) AS avg_variance,
                ROUND(AVG(actual_machining/NULLIF(standard_machining,0)),3) AS margin_ratio,
                ROUND(AVG(vendor_price),0) AS avg_price,
                ROUND(MIN(vendor_price),0) AS min_price,
                ROUND(MAX(vendor_price),0) AS max_price
            FROM history_log
            WHERE vendor_id=? AND difficulty_level IS NOT NULL
            GROUP BY difficulty_level
            ORDER BY difficulty_level
        """, conn, params=(vendor_id,))
    except Exception:
        return pd.DataFrame()
    finally:
        conn.close()

def search_similar_parts(volume_cm3, hole_count, setup_count,
                         exclude_id=None) -> pd.DataFrame:
    if not DB_PATH.exists():
        return pd.DataFrame()
    conn = sqlite3.connect(DB_PATH)
    try:
        exclude_clause = f"AND h.id != {exclude_id}" if exclude_id else ""
        return pd.read_sql(f"""
            SELECT h.id, h.created_at, h.part_name, h.part_no,
                   h.vendor_name, h.vendor_price, h.material_cost,
                   h.variance_pct, h.verdict,
                   h.volume_cm3, h.hole_count, h.setup_count, h.difficulty_level,
                   ABS(h.volume_cm3 - {volume_cm3}) AS vol_diff
            FROM history_log h
            WHERE h.volume_cm3 BETWEEN {volume_cm3*0.8} AND {volume_cm3*1.2}
              AND h.hole_count BETWEEN {hole_count-2} AND {hole_count+2}
              AND h.setup_count = {setup_count}
              AND h.volume_cm3 IS NOT NULL
              {exclude_clause}
            ORDER BY vol_diff ASC LIMIT 8
        """, conn)
    except Exception:
        return pd.DataFrame()
    finally:
        conn.close()


# ══════════════════════════════════════════════════════════════
# 엑셀 리포트 생성
# ══════════════════════════════════════════════════════════════
def _cs(ws, row, col, value, bold=False, bg=None, align="left",
        num_format=None, font_color="000000", size=10):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=font_color, name="맑은 고딕", size=size)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    t = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=t, right=t, top=t, bottom=t)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if num_format:
        cell.number_format = num_format
    return cell

def build_report_excel(part_name, part_no, vendor_name, material_code,
                       result: AuditResult, difficulty_level, estimated_hours,
                       similar_df, sim_df, audit_date) -> bytes:
    wb = Workbook()
    C_HDR = "1B2A4A"; C_SUB = "2563EB"
    C_G = "E8F5E9"; C_R = "FFEBEE"; C_O = "FFF3E0"; C_Y = "FFFDE7"; C_GR = "F5F5F5"
    vbg = {"경고 · 과다 청구": C_R, "주의 · 소폭 과다": C_O,
           "신뢰 · 적정": C_G, "관찰 · 소폭 저가": C_Y, "경고 · 저가 수주": C_R}

    # 시트1: 감사 결과
    ws1 = wb.active; ws1.title = "감사 결과"
    for c, w in zip("ABCD", [22,22,18,18]):
        ws1.column_dimensions[c].width = w
    ws1.row_dimensions[1].height = 32
    ws1.merge_cells("A1:D1")
    _cs(ws1,1,1,"가공 견적 역산 감사 리포트",True,C_HDR,"center","","FFFFFF",14)
    ws1.merge_cells("A2:D2")
    _cs(ws1,2,1,f"감사 일시: {audit_date}",False,C_GR)
    r=3
    for lbl,val in [("부품명",part_name),("형번",part_no or "—"),
                    ("협력사",vendor_name or "—"),("소재",material_code),
                    ("난이도",f"{difficulty_level}등급 ×{get_coefficient(difficulty_level)}"),
                    ("예상 가공시간",f"{estimated_hours:.1f} h")]:
        _cs(ws1,r,1,lbl,True,C_GR)
        ws1.merge_cells(f"B{r}:D{r}"); _cs(ws1,r,2,val); r+=1
    r+=1
    ws1.merge_cells(f"A{r}:D{r}")
    _cs(ws1,r,1,"역산 수치",True,C_SUB,"center","","FFFFFF"); r+=1
    for lbl,val,fmt in [
        ("협력사 단가",result.vendor_price,"#,##0 원"),
        ("재료비",result.material_cost,"#,##0 원"),
        ("후처리비",result.postprocess_cost,"#,##0 원"),
        ("실질 가공비",result.actual_machining,"#,##0 원"),
        ("적정 가공비",result.standard_machining,"#,##0 원"),
        ("오차율",result.variance_pct,'0.0 "%"'),
        ("임률",result.hourly_rate,"#,##0 원/h"),
    ]:
        _cs(ws1,r,1,lbl,True,C_GR)
        ws1.merge_cells(f"B{r}:D{r}"); _cs(ws1,r,2,val,num_format=fmt); r+=1
    r+=1
    bg_v = vbg.get(result.verdict, "FFFFFF")
    ws1.merge_cells(f"A{r}:D{r}"); ws1.row_dimensions[r].height = 24
    _cs(ws1,r,1,f"판정: {result.verdict} ({result.variance_pct:.1f}%)",True,bg_v,"center"); r+=1
    ws1.merge_cells(f"A{r}:D{r}")
    _cs(ws1,r,1,f"권장 액션: {result.action}",False,bg_v,"center"); r+=2
    ws1.merge_cells(f"A{r}:D{r}")
    _cs(ws1,r,1,"협상 근거",True,C_SUB,"center","","FFFFFF"); r+=1
    diff = result.actual_machining - result.standard_machining
    pct  = result.variance_pct
    if pct > 110:
        nego = (f"협력사 단가 {result.vendor_price:,.0f}원 기준 실질 가공비가 "
                f"적정 가공비({result.standard_machining:,.0f}원) 대비 {pct-100:.1f}% 초과.\n"
                f"과다 청구 추정액: 약 {diff:,.0f}원\n"
                f"적정 단가 제안: {result.vendor_price-diff:,.0f}원 수준")
    elif pct < 80:
        nego = (f"단가 {result.vendor_price:,.0f}원은 적정 가공비 대비 {100-pct:.1f}% 낮습니다.\n"
                f"품질 조건 및 원자재 규격을 명기하고 진행을 권장합니다.")
    else:
        nego = (f"단가 {result.vendor_price:,.0f}원은 오차율 {pct:.1f}%로 적정 범위입니다.\n"
                f"현재 단가로 진행해도 무방합니다.")
    ws1.merge_cells(f"A{r}:D{r+2}")
    c = ws1.cell(row=r, column=1, value=nego)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    t = Side(style="thin", color="CCCCCC")
    c.border = Border(left=t, right=t, top=t, bottom=t)
    ws1.row_dimensions[r].height = 60

    # 시트2: 유사 부품 비교
    ws2 = wb.create_sheet("유사 부품 비교")
    for c, w in zip("ABCDEFG", [14,18,14,16,14,12,12]):
        ws2.column_dimensions[c].width = w
    ws2.merge_cells("A1:G1")
    _cs(ws2,1,1,"유사 부품 이력 비교",True,C_HDR,"center","","FFFFFF",12)
    for ci, h in enumerate(["일시","부품명","형번","업체","단가(원)","오차율(%)","판정"], 1):
        _cs(ws2,2,ci,h,True,C_SUB,"center","","FFFFFF")
    if similar_df is not None and not similar_df.empty:
        for ri, (_, row) in enumerate(similar_df.iterrows(), 3):
            bg2 = vbg.get(str(row.get("verdict","")), "FFFFFF")
            for ci, (v, f) in enumerate(zip(
                [str(row.get("created_at",""))[:10], row.get("part_name",""),
                 row.get("part_no","") or "—", row.get("vendor_name","") or "—",
                 row.get("vendor_price",0), row.get("variance_pct",0), row.get("verdict","")],
                [None,None,None,None,"#,##0","0.0",None]), 1):
                _cs(ws2,ri,ci,v,num_format=f,bg=bg2 if ci==7 else "FFFFFF")
    else:
        ws2.merge_cells("A3:G3")
        _cs(ws2,3,1,"유사 부품 이력이 없습니다.",bg=C_GR,align="center")

    # 시트3: 시뮬레이션
    if sim_df is not None and not sim_df.empty:
        ws3 = wb.create_sheet("업체별 시뮬레이션")
        for c, w in zip("ABCDEF", [20,16,16,16,14,22]):
            ws3.column_dimensions[c].width = w
        ws3.merge_cells("A1:F1")
        _cs(ws3,1,1,f"업체별 예상 단가 — {part_name}",True,C_HDR,"center","","FFFFFF",12)
        for ci, h in enumerate(["업체명","예상 단가(원)","예상 가공비(원)","재료비(원)","예상 오차율(%)","비고"], 1):
            _cs(ws3,2,ci,h,True,C_SUB,"center","","FFFFFF")
        for ri, (_, row) in enumerate(sim_df.iterrows(), 3):
            pv = row.get("예상오차율", 0)
            bg3 = C_O if pv > 110 else (C_Y if pv < 80 else C_G)
            for ci, (v, f) in enumerate(zip(
                [row.get("업체명",""), row.get("예상단가",0), row.get("예상가공비",0),
                 row.get("재료비",0), pv, row.get("비고","")],
                [None,"#,##0","#,##0","#,##0","0.0",None]), 1):
                _cs(ws3,ri,ci,v,num_format=f,bg=bg3 if ci in (2,5) else "FFFFFF")

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════════
# 판정 UI
# ══════════════════════════════════════════════════════════════
def render_verdict(result: AuditResult):
    fn = {"green": st.success, "orange": st.warning,
          "yellow": st.warning, "red": st.error}.get(result.color, st.info)
    fn(f"**{result.verdict}** ({result.variance_pct:.1f}%) — {result.action}")

def render_cost_cards(result: AuditResult):
    c1, c2, c3 = st.columns(3)
    c1.metric("재료비", f"{result.material_cost:,.0f} 원")
    c2.metric("실질 가공비", f"{result.actual_machining:,.0f} 원")
    delta = result.actual_machining - result.standard_machining
    c3.metric("적정 가공비", f"{result.standard_machining:,.0f} 원",
              delta=f"{delta:+,.0f} 원", delta_color="inverse")

def render_similar_parts(volume_cm3, hole_count, setup_count,
                         current_vendor_price, saved_id=None):
    if not (volume_cm3 and hole_count is not None and setup_count):
        return pd.DataFrame()
    similar = search_similar_parts(volume_cm3, hole_count, setup_count, saved_id)
    if similar.empty:
        return pd.DataFrame()
    st.subheader("🔍 유사 부품 이력")
    st.caption(f"부피 {volume_cm3:.0f}cm³ ±20% / 홀 {hole_count}개 ±2 / 셋업 {setup_count}회 — 전 업체")
    col_map = {"created_at":"일시","part_name":"부품명","part_no":"형번",
               "vendor_name":"업체","vendor_price":"단가","variance_pct":"오차율(%)","verdict":"판정"}
    st.dataframe(
        similar[[c for c in col_map if c in similar.columns]].rename(columns=col_map),
        hide_index=True, use_container_width=True,
        column_config={"오차율(%)": st.column_config.NumberColumn(format="%.1f %%"),
                       "단가": st.column_config.NumberColumn(format="%,.0f 원")}
    )
    valid = similar[similar["vendor_price"].notna() & (similar["vendor_price"] > 0)]
    if not valid.empty:
        avg_p = valid["vendor_price"].mean()
        min_p = valid["vendor_price"].min()
        n     = len(valid)
        diff  = (current_vendor_price - avg_p) / avg_p * 100
        if abs(diff) >= 3:
            direction = "높습니다" if diff > 0 else "낮습니다"
            msg = (f"유사 부품 {n}건 평균 **{avg_p:,.0f}원** 대비 "
                   f"현재 견적 **{current_vendor_price:,.0f}원** — "
                   f"**{abs(diff):.1f}% {direction}** (이력 최저: {min_p:,.0f}원)")
            (st.error if diff > 10 else st.warning if diff > 3 else st.info)(f"💬 협상 근거: {msg}")
    return similar


# ══════════════════════════════════════════════════════════════
# 시뮬레이션 계산
# ══════════════════════════════════════════════════════════════
def calc_vendor_simulation(volume_cm3, material_code, form_key, loss_rate,
                            hourly_rate, estimated_hours, difficulty_coeff,
                            postprocess_cost, price_override=None) -> pd.DataFrame:
    try:
        mat = calc_material_cost(volume_cm3=volume_cm3, material_code=material_code,
                                  form=form_key, loss_rate_override=loss_rate,
                                  price_override=price_override)
    except Exception:
        return pd.DataFrame()
    mc  = mat["material_cost"]
    std = hourly_rate * estimated_hours * difficulty_coeff
    stats = load_vendor_stats()
    if stats.empty:
        return pd.DataFrame()
    rows = []
    for _, row in stats.iterrows():
        margin = float(row["avg_margin_ratio"]) if row["avg_margin_ratio"] else 1.0
        if margin <= 0: margin = 1.0
        total  = int(row["total"]) if row["total"] else 0
        est_m  = std * margin
        est_p  = mc + est_m + postprocess_cost
        est_v  = (est_m / std * 100) if std > 0 else 0
        note   = (f"이력 {total}건 기반 (평균 오차율 {row['avg_variance']}%)"
                  if total >= 3 else f"이력 {total}건 — 데이터 부족")
        rows.append({"업체명":row["name"],"예상단가":round(est_p),
                     "예상가공비":round(est_m),"재료비":round(mc),
                     "예상오차율":round(est_v,1),"이력건수":total,"비고":note})
    return pd.DataFrame(rows).sort_values("예상단가") if rows else pd.DataFrame()


# ══════════════════════════════════════════════════════════════
# 메인
# ══════════════════════════════════════════════════════════════
def main():
    st.set_page_config(page_title="가공 견적 AI 엔진 PRO v2.3",
                       page_icon="🔧", layout="wide")
    inject_css()
    init_db()

    if "show_add_vendor" not in st.session_state:
        st.session_state.show_add_vendor = False

    with st.sidebar:
        st.markdown("""
        <div style='padding:8px 0 16px'>
            <div style='font-size:18px;font-weight:700;color:#FFFFFF;letter-spacing:-.3px'>
                🔧 Quote AI
            </div>
            <div style='font-size:11px;color:#94A3B8;margin-top:2px'>
                가공 견적 역산 감사 v2.3
            </div>
        </div>
        """, unsafe_allow_html=True)

        menu = st.radio("메뉴", [
            "📋 역산 감사", "📦 일괄 업로드", "📈 업체 시뮬레이션",
            "🏢 업체 관리", "📊 업체 비교", "📁 감사 이력"
        ], label_visibility="collapsed")

        st.divider()

        mat_df      = load_material_master()
        mat_labels  = [f"{r['material_code']} — {r['material_name']}" for _, r in mat_df.iterrows()]
        sel_mat_lbl = st.selectbox("소재", mat_labels)
        sel_mat_code= mat_df["material_code"].tolist()[mat_labels.index(sel_mat_lbl)]
        mat_info    = mat_df[mat_df["material_code"] == sel_mat_code].iloc[0]

        form_type   = st.radio("원자재 형태", ["봉재","판재"], horizontal=True)
        form_key    = "bar" if form_type == "봉재" else "plate"
        default_loss= float(mat_info[f"default_loss_{form_key}"])
        loss_rate   = st.slider("로스율 (%)", 0, 30, int(default_loss*100)) / 100

        st.divider()
        hourly_rate  = st.number_input("임률 (원/h)", min_value=10000,
                                       max_value=200000, value=45000, step=1000)
        st.caption("시세 설정")
        use_override  = st.toggle("수동 입력")
        price_override= None
        if use_override:
            price_override = st.number_input("직접 입력 (원/kg)", min_value=0,
                                             value=4800, step=100)

        prices = get_market_prices()
        source = prices.get("source", "알 수 없음")
        updated = prices.get("updated_at")

        if source == "KOMIS API":
            st.success(f"✅ {source}")
        else:
            st.warning(f"⚠ {source}")

        # 시세 갱신 경고
        if updated:
            try:
                last_update = datetime.fromisoformat(updated)
                days_ago = (datetime.now() - last_update).days
                if days_ago >= 3:
                    st.error(f"⚠ 시세 {days_ago}일째 미갱신")
                else:
                    st.caption(f"갱신: {last_update.strftime('%m/%d %H:%M')}")
            except Exception:
                pass

        if st.button("🔄 시세 새로고침"):
            get_market_prices(force_refresh=True)
            st.rerun()

    if   menu == "📋 역산 감사":
        page_audit(mat_df, sel_mat_code, mat_info, form_key,
                   loss_rate, hourly_rate, price_override)
    elif menu == "📦 일괄 업로드":
        page_bulk(mat_df, sel_mat_code, mat_info, form_key,
                  loss_rate, hourly_rate, price_override)
    elif menu == "📈 업체 시뮬레이션":
        page_simulation(mat_df, sel_mat_code, mat_info, form_key,
                        loss_rate, hourly_rate, price_override)
    elif menu == "🏢 업체 관리":
        page_vendors()
    elif menu == "📊 업체 비교":
        page_compare()
    elif menu == "📁 감사 이력":
        page_history()


# ══════════════════════════════════════════════════════════════
# 페이지: 역산 감사
# ══════════════════════════════════════════════════════════════
def page_audit(mat_df, sel_mat_code, mat_info, form_key,
               loss_rate, hourly_rate, price_override):
    st.title("📋 역산 감사")
    col_left, col_right = st.columns([1, 1], gap="large")

    with col_left:
        st.subheader("1 · 형상 정보")
        step_file = st.file_uploader(
            "STEP 파일 업로드 (솔리드웍스에서 .step으로 내보내기 후 업로드)",
            type=["step","stp"],
            help="솔리드웍스 → 파일 → 다른 이름으로 저장 → STEP(.step) 선택"
        )
        part_name = part_no = step_fname = ""
        volume_cm3 = hole_count = setup_count = auto_diff = None

        if step_file:
            part_name  = Path(step_file.name).stem
            step_fname = step_file.name
            if is_step_available():
                with st.spinner("STEP 분석 중..."):
                    with tempfile.NamedTemporaryFile(suffix=".step", delete=False) as tmp:
                        tmp.write(step_file.read()); tmp_path = tmp.name
                    res = analyze_step_file(tmp_path, float(mat_info.get("machinability",1.0)))
                    os.unlink(tmp_path)
                if res["error"]:
                    st.error(f"분석 오류: {res['error']}")
                else:
                    bb = res["bounding_box"]
                    volume_cm3  = bb.get("volume_cm3", 0)
                    holes       = res["holes"]
                    hole_count  = len(holes)
                    setup_count = res["setups"]
                    auto_diff   = res["difficulty"]
                    st.success("✅ STEP 분석 완료")
                    c1,c2,c3 = st.columns(3)
                    c1.metric("X", f"{bb.get('x_mm',0):.1f} mm")
                    c2.metric("Y", f"{bb.get('y_mm',0):.1f} mm")
                    c3.metric("Z", f"{bb.get('z_mm',0):.1f} mm")
                    st.caption(f"홀: {hole_count}개 · 셋업 추정: {setup_count}회")
            else:
                st.info("💡 STEP 자동 분석은 로컬 PC(pythonOCC 설치)에서만 동작합니다. 아래에 직접 입력해 주세요.")

        part_name  = st.text_input("부품명", value=part_name)
        part_no    = st.text_input("형번", value=part_no, help="예: A001, BRKT-001")
        step_fname = st.text_input("도면 파일명", value=step_fname)

        col_v, col_h = st.columns(2)
        if volume_cm3 is None:
            volume_cm3 = col_v.number_input("바운딩 부피 (cm³)", min_value=0.1, value=100.0, step=10.0)
        else:
            col_v.number_input("바운딩 부피 (cm³) — 자동", value=volume_cm3, disabled=True)
        if hole_count is None:
            hole_count = col_h.number_input("홀 개수", min_value=0, value=0, step=1)

        if setup_count is None:
            setup_count = st.number_input("셋업 횟수", min_value=1, value=1, step=1)

        st.divider()
        st.subheader("2 · 난이도")
        if auto_diff:
            st.info(f"자동 추정: **{auto_diff['level']}등급** · {auto_diff['reason']}")
            default_idx = auto_diff["level"] - 1
        else:
            default_idx = 0
        selected_diff    = st.radio("난이도 등급", get_all_options(),
                                    index=default_idx, horizontal=True)
        difficulty_level = parse_level_from_option(selected_diff)
        difficulty_coeff = get_coefficient(difficulty_level)
        st.caption(f"계수: **{difficulty_coeff}** — {DIFFICULTY_TABLE[difficulty_level]['desc']}")

    with col_right:
        st.subheader("3 · 협력사 단가")
        vendors_df = load_vendors()
        if vendors_df.empty:
            st.warning("등록된 업체가 없습니다. [업체 관리]에서 먼저 등록해주세요.")
            vendor_id   = None
            vendor_name = st.text_input("업체명 (직접 입력)")
        else:
            vendor_options = ["직접 입력"] + vendors_df["name"].tolist()
            sel_vendor = st.selectbox("협력사 선택", vendor_options)
            if sel_vendor == "직접 입력":
                vendor_id   = None
                vendor_name = st.text_input("업체명")
            else:
                vrow        = vendors_df[vendors_df["name"] == sel_vendor].iloc[0]
                vendor_id   = int(vrow["id"])
                vendor_name = sel_vendor
                st.caption(f"담당: {vrow.get('contact','—') or '—'} · {vrow.get('phone','—') or '—'}")

        vendor_price     = st.number_input("협력사 단가 (원)", min_value=0, value=100000, step=1000)
        col_pp, col_eh   = st.columns(2)
        postprocess_cost = col_pp.number_input("후처리비 (원)", min_value=0, value=0, step=1000,
                                               help="도금·열처리 등")
        estimated_hours  = col_eh.number_input("가공시간 (h)", min_value=0.1, value=1.0, step=0.1)
        note             = st.text_input("메모 (선택)")

        st.divider()
        try:
            mat_result = calc_material_cost(volume_cm3=volume_cm3, material_code=sel_mat_code,
                                             form=form_key, loss_rate_override=loss_rate,
                                             price_override=price_override)
            st.metric("예상 재료비", f"{mat_result['material_cost']:,.0f} 원",
                      help=f"{mat_result['weight_kg']:.3f}kg × 로스 {mat_result['loss_rate']*100:.0f}% × {mat_result['price_used']:,.0f}원/kg")
        except Exception as e:
            st.warning(f"재료비 미리보기 오류: {e}")
            mat_result = None

        run_audit = st.button("🔍 감사 실행", type="primary", use_container_width=True)

    if run_audit:
        if not part_name:
            st.warning("부품명을 입력해주세요.")
        elif mat_result is None:
            st.error("재료비 계산 오류.")
        else:
            try:
                result = audit(vendor_price=vendor_price,
                               material_cost=mat_result["material_cost"],
                               hourly_rate=hourly_rate, estimated_hours=estimated_hours,
                               difficulty_coeff=difficulty_coeff,
                               postprocess_cost=postprocess_cost)
                st.divider()
                st.subheader("📊 감사 결과")
                render_verdict(result)
                render_cost_cards(result)

                with st.expander("상세 수치"):
                    st.dataframe(pd.DataFrame({
                        "항목": ["협력사 단가","재료비","후처리비","실질 가공비",
                                 "적정 가공비","오차율","임률","가공시간","난이도 계수"],
                        "값":   [f"{result.vendor_price:,.0f} 원",
                                 f"{result.material_cost:,.0f} 원",
                                 f"{result.postprocess_cost:,.0f} 원",
                                 f"{result.actual_machining:,.0f} 원",
                                 f"{result.standard_machining:,.0f} 원",
                                 f"{result.variance_pct:.1f} %",
                                 f"{result.hourly_rate:,.0f} 원/h",
                                 f"{result.estimated_hours:.1f} h",
                                 f"{result.difficulty_coeff}"]
                    }), hide_index=True)

                st.subheader("📈 재료비 시나리오")
                adj = st.slider("재료비 조정률 (%)", -30, 30, 0, 1)
                if adj != 0:
                    sc = audit_scenario(result, adj)
                    s1,s2,s3 = st.columns(3)
                    s1.metric("조정 재료비", f"{sc.material_cost:,.0f} 원", delta=f"{adj:+d}%")
                    s2.metric("조정 실질 가공비", f"{sc.actual_machining:,.0f} 원")
                    s3.metric("조정 오차율", f"{sc.variance_pct:.1f}%",
                              delta=f"{sc.variance_pct-result.variance_pct:+.1f}%",
                              delta_color="inverse")
                    render_verdict(sc)

                save_to_db(part_name=part_name, part_no=part_no, step_file=step_fname,
                           material_code=sel_mat_code, vendor_id=vendor_id,
                           vendor_name=vendor_name, result=result,
                           difficulty_level=difficulty_level, estimated_hours=estimated_hours,
                           volume_cm3=volume_cm3, hole_count=hole_count,
                           setup_count=setup_count, price_snapshot=mat_result["price_used"],
                           note=note)
                st.success("✅ 이력 저장 완료")

                try:
                    conn = sqlite3.connect(DB_PATH)
                    last_id = conn.execute("SELECT MAX(id) FROM history_log").fetchone()[0]
                    conn.close()
                except Exception:
                    last_id = None

                similar_df = render_similar_parts(
                    float(volume_cm3), int(hole_count), int(setup_count),
                    vendor_price, last_id)

                # 협상 근거 엑셀
                st.divider()
                st.subheader("📥 협상 근거 리포트")
                sim_df = calc_vendor_simulation(
                    float(volume_cm3), sel_mat_code, form_key, loss_rate,
                    hourly_rate, estimated_hours, difficulty_coeff,
                    postprocess_cost, price_override)
                excel_bytes = build_report_excel(
                    part_name, part_no or "", vendor_name or "", sel_mat_code,
                    result, difficulty_level, estimated_hours,
                    similar_df if isinstance(similar_df, pd.DataFrame) else pd.DataFrame(),
                    sim_df if not sim_df.empty else None,
                    datetime.now().strftime("%Y-%m-%d %H:%M"))
                st.download_button(
                    "📊 협상 근거 엑셀 다운로드",
                    data=excel_bytes,
                    file_name=f"협상근거_{part_name}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary", use_container_width=True)

            except Exception as e:
                st.error(f"감사 실행 오류: {e}")


# ══════════════════════════════════════════════════════════════
# 페이지: 일괄 업로드
# ══════════════════════════════════════════════════════════════
def page_bulk(mat_df, sel_mat_code, mat_info, form_key,
              loss_rate, hourly_rate, price_override):
    st.title("📦 일괄 업로드")
    st.caption("STEP 파일 여러 개 + 명세서(Excel/CSV)를 동시에 업로드하면 형번을 자동 매칭합니다.")

    col_l, col_r = st.columns(2, gap="large")

    with col_l:
        st.subheader("STEP 파일")
        st.caption("솔리드웍스 → 파일 → 다른 이름으로 저장 → STEP(.step)")
        step_files = st.file_uploader(
            "STEP 파일 업로드 (여러 개 선택 가능)",
            type=["step","stp"], accept_multiple_files=True,
            key="bulk_step"
        )
        if step_files:
            st.success(f"✅ {len(step_files)}개 파일 업로드됨")
            for f in step_files:
                st.caption(f"· {f.name}")

    with col_r:
        st.subheader("명세서 (선택)")
        st.caption("형번, 단가, 업체명 컬럼이 있는 Excel 또는 CSV")
        spec_file = st.file_uploader(
            "명세서 업로드",
            type=["xlsx","xls","csv"],
            key="bulk_spec"
        )
        spec_df = None
        if spec_file:
            try:
                if spec_file.name.endswith(".csv"):
                    spec_df = pd.read_csv(spec_file)
                else:
                    spec_df = pd.read_excel(spec_file)
                st.success(f"✅ {len(spec_df)}행 파싱 완료")
                st.dataframe(spec_df.head(3), hide_index=True, use_container_width=True)
            except Exception as e:
                st.error(f"명세서 읽기 오류: {e}")

    if not step_files:
        st.info("👆 STEP 파일을 먼저 업로드해주세요.")
        return

    st.divider()
    st.subheader("형번 자동 매칭")

    # 파일명에서 형번 추출 (영숫자+하이픈 패턴)
    import re
    rows = []
    for f in step_files:
        stem  = Path(f.name).stem
        # 파일명에서 형번 패턴 추출 (예: BRKT-MAIN-A2301 → A2301, A-2301 등)
        match = re.search(r'([A-Z]{1,4}[-_]?\d{3,6})', stem.upper())
        part_no_guess = match.group(1).replace("_","-") if match else ""

        # 명세서에서 매칭 시도
        vendor_from_spec = ""
        price_from_spec  = 0
        matched = False
        if spec_df is not None:
            # 형번 컬럼 자동 감지
            no_col = next((c for c in spec_df.columns
                           if any(k in c.lower() for k in ["형번","part_no","partno","no"])), None)
            price_col = next((c for c in spec_df.columns
                              if any(k in c.lower() for k in ["단가","price","금액"])), None)
            vendor_col = next((c for c in spec_df.columns
                               if any(k in c.lower() for k in ["업체","vendor","회사","company"])), None)
            if no_col and part_no_guess:
                hit = spec_df[spec_df[no_col].astype(str).str.contains(
                    re.escape(part_no_guess), case=False, na=False)]
                if not hit.empty:
                    row = hit.iloc[0]
                    vendor_from_spec = str(row[vendor_col]) if vendor_col else ""
                    price_from_spec  = float(row[price_col]) if price_col else 0
                    matched = True

        rows.append({
            "파일명":   f.name,
            "추출 형번": part_no_guess,
            "부품명":   stem,
            "업체":     vendor_from_spec,
            "단가":     price_from_spec,
            "상태":     "✅ 매칭" if matched else ("⚠ 형번만" if part_no_guess else "❌ 미매칭"),
            "_file":    f,
        })

    match_df = pd.DataFrame(rows)
    matched_cnt  = len(match_df[match_df["상태"] == "✅ 매칭"])
    partial_cnt  = len(match_df[match_df["상태"] == "⚠ 형번만"])
    unmatch_cnt  = len(match_df[match_df["상태"] == "❌ 미매칭"])

    c1,c2,c3,c4 = st.columns(4)
    c1.metric("전체", f"{len(match_df)}건")
    c2.metric("✅ 매칭", f"{matched_cnt}건")
    c3.metric("⚠ 형번만", f"{partial_cnt}건")
    c4.metric("❌ 미매칭", f"{unmatch_cnt}건")

    # 편집 가능한 매칭 테이블
    st.caption("아래에서 형번·업체·단가를 직접 수정할 수 있습니다.")
    vendors_list = [""] + load_vendors()["name"].tolist() if not load_vendors().empty else [""]

    edited_rows = []
    for i, row in match_df.iterrows():
        with st.expander(f"{row['파일명']}  —  {row['상태']}", expanded=(row['상태'] != '✅ 매칭')):
            ec1, ec2, ec3, ec4 = st.columns([2,2,2,2])
            pno   = ec1.text_input("형번",  value=row["추출 형번"], key=f"bpno_{i}")
            pname = ec2.text_input("부품명", value=row["부품명"], key=f"bpname_{i}")
            vend  = ec3.text_input("업체명", value=row["업체"], key=f"bvend_{i}")
            price = ec4.number_input("단가(원)", value=float(row["단가"]),
                                     min_value=0.0, step=1000.0, key=f"bprice_{i}")
            edited_rows.append({**row, "추출 형번":pno, "부품명":pname,
                                 "업체":vend, "단가":price})

    st.divider()

    # 난이도 + 가공시간 공통 설정
    st.subheader("공통 설정")
    st.caption("아래 설정은 일괄 적용됩니다. 개별 조정이 필요하면 역산 감사 페이지를 이용하세요.")
    bc1, bc2, bc3 = st.columns(3)
    bulk_diff_s  = bc1.selectbox("난이도 (공통)", get_all_options(), key="bulk_diff")
    bulk_diff_lv = parse_level_from_option(bulk_diff_s)
    bulk_coeff   = get_coefficient(bulk_diff_lv)
    bulk_hours   = bc2.number_input("가공시간 (h, 공통)", min_value=0.1, value=1.0,
                                    step=0.1, key="bulk_hours")
    bulk_pp      = bc3.number_input("후처리비 (원, 공통)", min_value=0, value=0,
                                    step=1000, key="bulk_pp")

    run_bulk = st.button("🚀 일괄 감사 실행", type="primary", use_container_width=True)

    if run_bulk:
        if not any(r["단가"] > 0 for r in edited_rows):
            st.warning("단가가 입력된 항목이 없습니다.")
            return

        results = []
        progress = st.progress(0)
        for idx, row in enumerate(edited_rows):
            progress.progress((idx+1) / len(edited_rows))
            if not row["단가"] or float(row["단가"]) <= 0:
                continue
            try:
                mat = calc_material_cost(volume_cm3=100.0, material_code=sel_mat_code,
                                          form=form_key, loss_rate_override=loss_rate,
                                          price_override=price_override)
                res = audit(vendor_price=float(row["단가"]),
                            material_cost=mat["material_cost"],
                            hourly_rate=hourly_rate, estimated_hours=bulk_hours,
                            difficulty_coeff=bulk_coeff, postprocess_cost=bulk_pp)

                # 업체 ID 조회
                vdf = load_vendors()
                vrow = vdf[vdf["name"] == row["업체"]] if row["업체"] else pd.DataFrame()
                vid  = int(vrow.iloc[0]["id"]) if not vrow.empty else None

                save_to_db(part_name=row["부품명"], part_no=row["추출 형번"],
                           step_file=row["파일명"], material_code=sel_mat_code,
                           vendor_id=vid, vendor_name=row["업체"],
                           result=res, difficulty_level=bulk_diff_lv,
                           estimated_hours=bulk_hours,
                           price_snapshot=mat["price_used"])

                results.append({"형번": row["추출 형번"], "부품명": row["부품명"],
                                 "업체": row["업체"], "단가": int(row["단가"]),
                                 "오차율(%)": round(res.variance_pct, 1),
                                 "판정": res.verdict})
            except Exception as e:
                results.append({"형번": row["추출 형번"], "부품명": row["부품명"],
                                 "업체": row["업체"], "단가": int(row["단가"]),
                                 "오차율(%)": 0, "판정": f"오류: {e}"})

        progress.empty()
        st.success(f"✅ {len(results)}건 감사 완료 · 이력 저장됨")

        res_df = pd.DataFrame(results)
        warn_cnt = len(res_df[res_df["오차율(%)"] > 110]) + len(res_df[res_df["오차율(%)"] < 80])
        ok_cnt   = len(res_df[(res_df["오차율(%)"] >= 90) & (res_df["오차율(%)"] <= 110)])
        sc1,sc2,sc3 = st.columns(3)
        sc1.metric("감사 완료", f"{len(results)}건")
        sc2.metric("✅ 신뢰·적정", f"{ok_cnt}건")
        sc3.metric("⚠ 주의·경고", f"{warn_cnt}건")

        st.dataframe(res_df, hide_index=True, use_container_width=True,
                     column_config={"오차율(%)": st.column_config.NumberColumn(format="%.1f %%"),
                                    "단가": st.column_config.NumberColumn(format="%,.0f 원")})

        csv = res_df.to_csv(index=False, encoding="utf-8-sig")
        st.download_button("📥 결과 CSV 다운로드", data=csv,
                           file_name=f"일괄감사_{datetime.now().strftime('%Y%m%d')}.csv",
                           mime="text/csv")


# ══════════════════════════════════════════════════════════════
# 페이지: 업체 시뮬레이션
# ══════════════════════════════════════════════════════════════
def page_simulation(mat_df, sel_mat_code, mat_info, form_key,
                    loss_rate, hourly_rate, price_override):
    st.title("📈 업체별 예상 단가 시뮬레이션")
    st.caption("신규 부품 발주 전, 등록 업체별 과거 마진 패턴 기반으로 예상 단가를 추정합니다.")

    col_l, col_r = st.columns([1,1], gap="large")

    with col_l:
        st.subheader("부품 사양")
        sim_name   = st.text_input("부품명 (참고용)", key="sim_name")
        c1, c2     = st.columns(2)
        sim_vol    = c1.number_input("바운딩 부피 (cm³)", min_value=0.1, value=100.0, step=10.0, key="sim_vol")
        sim_hole   = c2.number_input("홀 개수", min_value=0, value=0, step=1, key="sim_hole")
        sim_setup  = st.number_input("셋업 횟수", min_value=1, value=1, step=1, key="sim_setup")
        sim_pp     = st.number_input("후처리비 (원)", min_value=0, value=0, step=1000, key="sim_pp")
        sim_diff_s = st.radio("난이도", get_all_options(), horizontal=True, key="sim_diff")
        sim_diff_lv= parse_level_from_option(sim_diff_s)
        sim_coeff  = get_coefficient(sim_diff_lv)
        sim_hours  = st.number_input("예상 가공시간 (h)", min_value=0.1, value=1.0, step=0.1, key="sim_hours")
        run_sim    = st.button("🔍 시뮬레이션 실행", type="primary", use_container_width=True)

    with col_r:
        st.subheader("예상 단가 순위")
        if run_sim:
            sim_df = calc_vendor_simulation(
                float(sim_vol), sel_mat_code, form_key, loss_rate,
                hourly_rate, sim_hours, sim_coeff, sim_pp, price_override)

            if sim_df.empty:
                st.info("등록된 업체가 없거나 이력이 부족합니다.")
            else:
                for rank, (_, row) in enumerate(sim_df.iterrows(), 1):
                    pct   = row["예상오차율"]
                    price = row["예상단가"]
                    total = row["이력건수"]
                    icon  = "🔴" if pct > 110 else ("🔵" if pct < 80 else "🟢")
                    stars = "★★★" if total >= 5 else ("★★☆" if total >= 3 else "★☆☆")
                    st.metric(f"{rank}위 {icon} {row['업체명']} {stars}",
                              f"{price:,.0f} 원",
                              delta=f"예상 오차율 {pct:.1f}%",
                              delta_color="inverse")

                st.divider()
                st.dataframe(
                    sim_df[["업체명","예상단가","예상가공비","재료비","예상오차율","이력건수","비고"]],
                    hide_index=True, use_container_width=True,
                    column_config={
                        "예상단가": st.column_config.NumberColumn("예상 단가(원)", format="%,.0f"),
                        "예상가공비": st.column_config.NumberColumn("예상 가공비(원)", format="%,.0f"),
                        "재료비": st.column_config.NumberColumn("재료비(원)", format="%,.0f"),
                        "예상오차율": st.column_config.NumberColumn("예상 오차율(%)", format="%.1f %%"),
                        "이력건수": st.column_config.NumberColumn("이력 건수"),
                    }
                )
                st.caption("★★★ 이력 5건↑  ★★☆ 3~4건  ★☆☆ 1~2건 (신뢰도 낮음)")

                # 엑셀 다운로드
                try:
                    dummy_result = audit(
                        vendor_price=float(sim_df["예상단가"].iloc[0]),
                        material_cost=float(sim_df["재료비"].iloc[0]),
                        hourly_rate=hourly_rate, estimated_hours=sim_hours,
                        difficulty_coeff=sim_coeff, postprocess_cost=sim_pp)
                    excel_bytes = build_report_excel(
                        sim_name or "신규부품", "", "시뮬레이션", sel_mat_code,
                        dummy_result, sim_diff_lv, sim_hours,
                        pd.DataFrame(), sim_df,
                        datetime.now().strftime("%Y-%m-%d %H:%M"))
                    st.download_button(
                        "📥 시뮬레이션 결과 엑셀",
                        data=excel_bytes,
                        file_name=f"시뮬_{sim_name or '신규'}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                except Exception:
                    pass


# ══════════════════════════════════════════════════════════════
# 페이지: 업체 관리
# ══════════════════════════════════════════════════════════════
def page_vendors():
    st.title("🏢 업체 관리")

    col_t, col_b = st.columns([4,1])
    with col_b:
        if st.button("➕ 업체 추가", use_container_width=True):
            st.session_state.show_add_vendor = not st.session_state.show_add_vendor

    if st.session_state.show_add_vendor:
        with st.container(border=True):
            st.subheader("신규 업체 등록")
            with st.form("add_vendor_form", clear_on_submit=True):
                c1, c2 = st.columns(2)
                name    = c1.text_input("업체명 *", placeholder="필수 입력")
                contact = c2.text_input("담당자명")
                phone   = c1.text_input("연락처", placeholder="010-0000-0000")
                email   = c2.text_input("이메일")
                note    = st.text_area("메모", height=80)
                cs, cc  = st.columns(2)
                submitted = cs.form_submit_button("✅ 등록", type="primary", use_container_width=True)
                cancelled = cc.form_submit_button("취소", use_container_width=True)
            if submitted:
                if not name:
                    st.error("업체명은 필수입니다.")
                else:
                    add_vendor(name, contact, phone, email, note)
                    st.success(f"'{name}' 등록 완료")
                    st.session_state.show_add_vendor = False
                    st.rerun()
            if cancelled:
                st.session_state.show_add_vendor = False
                st.rerun()

    st.divider()
    vendors_df = load_vendors()
    if vendors_df.empty:
        st.info("등록된 업체가 없습니다. [업체 추가] 버튼을 눌러 등록해주세요.")
        return

    stats_df = load_vendor_stats()

    for _, row in vendors_df.iterrows():
        vid  = int(row["id"])
        stat = stats_df[stats_df["id"] == vid] if not stats_df.empty else pd.DataFrame()

        with st.expander(f"**{row['name']}**  ·  {row.get('contact','') or '담당자 미등록'}"):
            c1,c2,c3 = st.columns(3)
            c1.write(f"📞 {row.get('phone','—') or '—'}")
            c2.write(f"✉️ {row.get('email','—') or '—'}")
            c3.write(f"📝 {row.get('note','—') or '—'}")

            if not stat.empty and int(stat.iloc[0]["total"]) > 0:
                s = stat.iloc[0]
                st.divider()
                m1,m2,m3,m4 = st.columns(4)
                m1.metric("감사 건수",   f"{int(s['total'])}건")
                m2.metric("평균 오차율", f"{s['avg_variance']}%")
                m3.metric("발주 건수",   f"{int(s['ordered_count'])}건")
                m4.metric("재료비 비중", f"{s['mat_ratio']}%")

                # 패턴 분석 (5건 이상)
                if int(s["total"]) >= 5:
                    avg_v = float(s["avg_variance"])
                    msg = ("⚠ 오차율 높음 — 항목별 소명 요청 검토" if avg_v > 115
                           else "🔶 소폭 과다 — 단가 재협의 여지" if avg_v > 105
                           else "✅ 전반적 적정 수준" if avg_v >= 90
                           else "🔵 저가 경향 — 품질 조건 명기 권장")
                    st.info(f"{msg} (평균 {avg_v:.1f}%)")

                    # 난이도별 마진 패턴
                    pattern_df = load_vendor_pattern(vid)
                    if not pattern_df.empty:
                        with st.expander("📊 난이도별 마진 패턴 분석"):
                            diff_labels = {1:"단순",2:"보통",3:"복잡",4:"특수"}
                            for _, pr in pattern_df.iterrows():
                                lv   = int(pr["difficulty_level"])
                                lbl  = diff_labels.get(lv, str(lv))
                                cnt  = int(pr["cnt"])
                                avg  = float(pr["avg_variance"])
                                mr   = float(pr["margin_ratio"])
                                icon = "🔴" if avg > 110 else ("🟢" if avg >= 90 else "🔵")
                                st.write(
                                    f"{icon} **{lv}등급 · {lbl}** — "
                                    f"이력 {cnt}건 · 평균 오차율 {avg:.1f}% · "
                                    f"마진 계수 {mr:.2f}x"
                                )
                                if mr > 1.15:
                                    st.caption(f"  → 이 업체는 **{lbl}** 부품에서 마진을 집중시키는 경향이 있습니다.")

                hist = load_vendor_history(vid)
                if not hist.empty:
                    st.dataframe(
                        hist[["created_at","part_name","part_no","vendor_price",
                              "variance_pct","verdict"]].rename(columns={
                            "created_at":"일시","part_name":"부품명","part_no":"형번",
                            "vendor_price":"단가","variance_pct":"오차율(%)","verdict":"판정"}),
                        hide_index=True, use_container_width=True,
                        column_config={"오차율(%)": st.column_config.NumberColumn(format="%.1f %%"),
                                       "단가": st.column_config.NumberColumn(format="%,.0f 원")})

            st.divider()
            ec1, _, ec3 = st.columns([2,2,1])
            with ec1:
                with st.popover("✏️ 정보 수정"):
                    with st.form(f"edit_{vid}"):
                        nname    = st.text_input("업체명",  value=row["name"])
                        ncontact = st.text_input("담당자",  value=row.get("contact","") or "")
                        nphone   = st.text_input("연락처",  value=row.get("phone","") or "")
                        nemail   = st.text_input("이메일",  value=row.get("email","") or "")
                        nnote    = st.text_area("메모",     value=row.get("note","") or "")
                        if st.form_submit_button("저장"):
                            update_vendor(vid, nname, ncontact, nphone, nemail, nnote)
                            st.success("수정 완료"); st.rerun()
            with ec3:
                if st.button("🗑 비활성화", key=f"del_{vid}"):
                    if vendor_has_history(vid):
                        st.warning("이력 유지 후 비활성화 처리됩니다.")
                    deactivate_vendor(vid); st.rerun()


# ══════════════════════════════════════════════════════════════
# 페이지: 업체 비교
# ══════════════════════════════════════════════════════════════
def page_compare():
    st.title("📊 업체 비교")
    stats_df = load_vendor_stats()
    if stats_df.empty or stats_df["total"].sum() == 0:
        st.info("감사 이력이 없습니다. 역산 감사를 먼저 실행해주세요.")
        return

    stats_df = stats_df[stats_df["total"] > 0].copy()
    cols = st.columns(max(len(stats_df), 1))
    for i, (_, row) in enumerate(stats_df.iterrows()):
        avg_v = float(row["avg_variance"]) if row["avg_variance"] else 0
        icon  = "🟢" if 90 <= avg_v <= 110 else ("🔴" if avg_v > 110 else "🔵")
        with cols[i]:
            st.metric(f"{icon} {row['name']}", f"평균 {avg_v:.1f}%", f"총 {int(row['total'])}건")

    st.divider()
    st.dataframe(
        stats_df.rename(columns={"name":"업체명","total":"건수",
            "avg_variance":"평균오차율(%)","min_variance":"최저",
            "max_variance":"최고","ordered_count":"발주건","mat_ratio":"재료비비중(%)"}
        )[["업체명","건수","평균오차율(%)","최저","최고","발주건","재료비비중(%)"]],
        hide_index=True, use_container_width=True)

    st.subheader("형번별 단가 비교")
    conn = sqlite3.connect(DB_PATH)
    try:
        cross_df = pd.read_sql("""
            SELECT h.part_no, h.part_name, v.name AS vendor_name,
                   h.vendor_price, h.variance_pct, h.verdict, h.created_at
            FROM history_log h JOIN vendors v ON h.vendor_id=v.id
            WHERE h.part_no IS NOT NULL AND h.part_no != ''
            ORDER BY h.part_no, h.created_at DESC
        """, conn)
    except Exception:
        cross_df = pd.DataFrame()
    finally:
        conn.close()

    if cross_df.empty:
        st.info("형번이 입력된 이력이 없습니다.")
    else:
        sel_part = st.selectbox("형번 선택", cross_df["part_no"].unique().tolist())
        st.dataframe(
            cross_df[cross_df["part_no"] == sel_part][
                ["vendor_name","vendor_price","variance_pct","verdict","created_at"]
            ].rename(columns={"vendor_name":"업체명","vendor_price":"단가",
                               "variance_pct":"오차율(%)","verdict":"판정","created_at":"일시"}),
            hide_index=True, use_container_width=True,
            column_config={"오차율(%)": st.column_config.NumberColumn(format="%.1f %%"),
                           "단가": st.column_config.NumberColumn(format="%,.0f 원")})


# ══════════════════════════════════════════════════════════════
# 페이지: 감사 이력
# ══════════════════════════════════════════════════════════════
def page_history():
    st.title("📁 감사 이력")
    df = load_history()
    if df.empty:
        st.info("저장된 이력이 없습니다.")
        return

    with st.expander("🔎 검색 필터", expanded=True):
        r1c1, r1c2, r1c3 = st.columns(3)
        f_verdict = r1c1.multiselect("판정", df["verdict"].dropna().unique().tolist())
        f_vendor  = r1c2.multiselect("업체", df["vendor_name"].dropna().unique().tolist())
        f_text    = r1c3.text_input("통합 검색", placeholder="부품명 / 형번 / 메모")
        r2c1, r2c2 = st.columns(2)
        df["created_at"] = pd.to_datetime(df["created_at"])
        min_d = df["created_at"].min().date()
        max_d = df["created_at"].max().date()
        date_range = r2c1.date_input("감사 일자", value=(min_d, max_d),
                                     min_value=min_d, max_value=max_d)
        v_min = float(df["variance_pct"].min()) if df["variance_pct"].notna().any() else 0.0
        v_max = float(df["variance_pct"].max()) if df["variance_pct"].notna().any() else 200.0
        var_range = r2c2.slider("오차율 범위 (%)", 0.0, 300.0,
                                (max(0.0, v_min), min(300.0, v_max)), 1.0)

    if f_verdict: df = df[df["verdict"].isin(f_verdict)]
    if f_vendor:  df = df[df["vendor_name"].isin(f_vendor)]
    if f_text:
        mask = (df["part_name"].str.contains(f_text, case=False, na=False) |
                df["part_no"].str.contains(f_text, case=False, na=False) |
                df["note"].str.contains(f_text, case=False, na=False))
        df = df[mask]
    if isinstance(date_range, (list, tuple)) and len(date_range) == 2:
        df = df[(df["created_at"] >= pd.Timestamp(date_range[0])) &
                (df["created_at"] < pd.Timestamp(date_range[1]) + pd.Timedelta(days=1))]
    df = df[df["variance_pct"].isna() |
            ((df["variance_pct"] >= var_range[0]) & (df["variance_pct"] <= var_range[1]))]

    col_map = {"id":"ID","created_at":"일시","part_name":"부품명","part_no":"형번",
               "vendor_name":"업체","vendor_price":"협력사단가","material_cost":"재료비",
               "actual_machining":"실질가공비","standard_machining":"적정가공비",
               "variance_pct":"오차율(%)","verdict":"판정","difficulty_level":"난이도",
               "volume_cm3":"부피(cm³)","hole_count":"홀","setup_count":"셋업","note":"메모"}
    show = [c for c in col_map if c in df.columns]
    df_show = df[show].rename(columns=col_map)

    st.dataframe(df_show, hide_index=True, use_container_width=True,
                 column_config={
                     "오차율(%)":  st.column_config.NumberColumn(format="%.1f %%"),
                     "협력사단가": st.column_config.NumberColumn(format="%,.0f 원"),
                     "재료비":     st.column_config.NumberColumn(format="%,.0f 원"),
                     "실질가공비": st.column_config.NumberColumn(format="%,.0f 원"),
                     "적정가공비": st.column_config.NumberColumn(format="%,.0f 원"),
                     "부피(cm³)":  st.column_config.NumberColumn(format="%.1f"),
                 })
    st.caption(f"전체 {len(load_history())}건 · 표시 {len(df_show)}건")
    st.divider()
    csv = df_show.to_csv(index=False, encoding="utf-8-sig")
    st.download_button("📥 CSV 다운로드", data=csv,
                       file_name=f"감사이력_{datetime.now().strftime('%Y%m%d')}.csv",
                       mime="text/csv")


if __name__ == "__main__":
    main()
